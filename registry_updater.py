"""
Functions to update the customer registry with data from inspection files.
"""
import os
from datetime import datetime

import pandas as pd
import openpyxl
from typing import List, Dict, Any, Tuple, Optional, Union

# Import logging
from logging_config import get_logger
logger = get_logger()

# Import from config instead of main
from config import OUTPUT_FILE

def update_customer_registry(registry_path: str, inspection_data: List[Dict[str, Any]]) -> Tuple[
    str, List[Dict[str, Any]]]:
    """Update the customer registry with data from inspection files.

    Args:
        registry_path: Path to the customer registry Excel file
        inspection_data: List of dictionaries containing inspection data

    Returns:
        Tuple containing the output file path and a list of processing results
    """
    logger.info(f"Updating customer registry: {os.path.basename(registry_path)}")
    logger.debug(f"Number of inspection files to process: {len(inspection_data)}")

    try:
        # Check that registry path exists
        if not registry_path or not os.path.exists(registry_path):
            logger.error(f"Customer registry file not found: {registry_path}")
            return "", [{
                "status": "Error",
                "message": f"Customer registry file not found: {registry_path}",
                "filename": "N/A"
            }]

        # Load the customer registry Excel file
        logger.debug(f"Loading workbook: {registry_path}")
        registry_wb = openpyxl.load_workbook(registry_path)
        logger.debug(f"Workbook loaded. Available sheets: {registry_wb.sheetnames}")

        # Check if the required sheet exists
        if "Kaikki" not in registry_wb.sheetnames:
            logger.error("Required sheet 'Kaikki' not found in registry file")
            return "", [{
                "status": "Error",
                "message": f"Required sheet 'Kaikki' not found in registry file",
                "filename": "N/A"
            }]

        registry_sheet = registry_wb["Kaikki"]
        logger.debug(f"Using sheet 'Kaikki' with dimensions: {registry_sheet.dimensions}")

        # Get existing data as dataframe
        logger.debug("Converting sheet data to DataFrame")
        data = []
        header = None

        # Read the sheet data
        logger.debug("Reading sheet data")
        row_count = 0
        for i, row in enumerate(registry_sheet.iter_rows(values_only=True)):
            if i == 0:
                header = list(row)
                logger.debug(f"Header row: {header}")
            else:
                data.append(dict(zip(header, row)))
                row_count += 1

        logger.debug(f"Read {row_count} data rows from registry")

        df_existing = pd.DataFrame(data)
        logger.debug(f"Created DataFrame with shape: {df_existing.shape}")

        # Process each inspection data record
        logger.info("Processing inspection data records")
        processed_records = _process_inspection_data(df_existing, inspection_data)
        logger.debug(f"Processed {len(processed_records)} records")

        # Added check for empty DataFrame
        if df_existing.empty:
            logger.error("DataFrame is empty after processing. No data to save.")
            return "", [{
                "status": "Error",
                "message": "No data to save in registry. DataFrame is empty after processing.",
                "filename": "N/A"
            }]

        # Save the updated registry
        logger.info("Saving updated registry")
        _save_updated_registry_with_full_preservation(registry_sheet, df_existing)
        logger.info(f"Registry saved to: {OUTPUT_FILE}")

        return OUTPUT_FILE, processed_records

    except Exception as e:
        logger.error(f"Error updating customer registry: {e}", exc_info=True)
        return "", [{
            "status": "Error",
            "message": f"Failed to update registry: {str(e)}",
            "filename": "N/A"
        }]

def _process_inspection_data(df_existing: pd.DataFrame, inspection_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Process inspection data and update the existing dataframe.

    Args:
        df_existing: DataFrame containing existing registry data
        inspection_data: List of dictionaries containing inspection data

    Returns:
        List of dictionaries containing processing results
    """
    logger.debug("Processing inspection data")
    processed_records = []

    # Process each inspection data record
    for data in inspection_data:
        filename = data.get("filename", "unknown")
        logger.debug(f"Processing inspection data for file: {filename}")

        if "error" in data:
            logger.warning(f"Error in inspection data: {data['error']}")
            processed_records.append({
                "status": "Error",
                "message": data["error"],
                "filename": filename
            })
            continue

        # Check if the device already exists in the registry
        logger.debug("Checking if device already exists in registry")
        found_idx = _find_existing_device(df_existing, data)

        if found_idx is not None:
            logger.debug(f"Device found in registry at index {found_idx}")
        else:
            logger.debug("Device not found in registry, will add new record")

        # Format dates for display
        inspection_date_formatted = None
        if data.get("inspection_date"):
            inspection_date_formatted = data["inspection_date"].strftime("%d.%m.%Y")

        next_inspection_date_formatted = None
        if data.get("next_inspection_date"):
            next_inspection_date_formatted = data["next_inspection_date"].strftime("%d.%m.%Y")

        kympitys_date_formatted = None
        if data.get("kympitys_date"):
            kympitys_date_formatted = data["kympitys_date"].strftime("%d.%m.%Y")

        # Prepare new data for update or insertion
        logger.debug("Preparing data for update/insertion")
        new_data = {
            "Aktiivinen": True,
            "Tilaaja": data.get("owner"),  # Now from Y-AB:6 (Työmaa)
            "Tilaajan laite": None,  # Leave empty as specified
            "Laitteen nimi": data.get("model"),  # From F-0:11
            "Laitteen sarjanumero": data.get("serial_number"),  # From G-0:12
            "Tarkastettu": inspection_date_formatted,  # From AD-AG:6
            "Seuraava tarkastus": next_inspection_date_formatted,  # Calculated
            "Seuraava kympitys": kympitys_date_formatted,  # From AB-AD:58 (year) and Y-Z:58 (month)
            "Huollettu/korjattu": None,  # Leave empty as specified
            "Työmaa": data.get("owner"),  # Same as Tilaaja
            "Lisätieto": None,  # Leave empty as specified
            "Verkkolaskutus1": None,  # Leave empty as specified
            "Verkkolaskutus2": None,  # Leave empty as specified
            "Maksuehto": None,  # Leave empty as specified
            "Tarkastuspöytäkirja": filename  # As specified
        }

        # Log the data we're using
        for key, value in new_data.items():
            logger.debug(f"  {key}: {value}")

        if found_idx is not None:
            # Update existing record
            logger.debug("Updating existing record")
            for key, value in new_data.items():
                if key in df_existing.columns:
                    df_existing.at[found_idx, key] = value
                    logger.debug(f"Updated field '{key}' with value '{value}'")
            status = "Updated"
        else:
            # Create a new row with all columns from the existing dataframe
            logger.debug("Creating new record")
            new_row = {}
            for col in df_existing.columns:
                new_row[col] = None

            # Update with the data we have
            for key, value in new_data.items():
                if key in new_row:
                    new_row[key] = value

            # Append the new row
            logger.debug("Appending new row to DataFrame")
            df_existing = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
            logger.debug(f"DataFrame shape after append: {df_existing.shape}")
            status = "Added"

        processed_records.append({
            "status": status,
            "serial_number": data.get("serial_number"),
            "model": data.get("model"),
            "owner": data.get("owner"),
            "filename": filename
        })
        logger.info(
            f"Device processed - Status: {status}, Serial: {data.get('serial_number')}, Model: {data.get('model')}")

    return processed_records

def _find_existing_device(df: pd.DataFrame, data: Dict[str, Any]) -> Optional[int]:
    """Find if a device already exists in the registry.

    Args:
        df: DataFrame containing existing registry data
        data: Dictionary containing inspection data

    Returns:
        int: Index of the existing device or None if not found
    """
    serial_to_find = str(data.get("serial_number", "")) if data.get("serial_number") is not None else ""
    model_to_find = str(data.get("model", "")) if data.get("model") is not None else ""

    logger.debug(f"Searching for device - Serial: {serial_to_find}, Model: {model_to_find}")

    if not serial_to_find:
        logger.debug("No serial number provided, device will be considered new")
        return None

    for idx, row in df.iterrows():
        # Convert everything to string for safer comparison
        row_serial = str(row.get("Laitteen sarjanumero", "")) if pd.notna(
            row.get("Laitteen sarjanumero", "")) else ""
        row_model = str(row.get("Laitteen nimi", "")) if pd.notna(row.get("Laitteen nimi", "")) else ""

        logger.debug(f"Comparing with registry row {idx} - Serial: '{row_serial}', Model: '{row_model}'")

        if row_serial and serial_to_find and row_serial == serial_to_find:
            logger.debug("Serial number match found")

            if not row_model or not model_to_find or row_model == model_to_find:
                logger.debug(f"Device found at index {idx}")
                return idx
            else:
                logger.debug("Serial number matched but model didn't match")

    logger.debug("Device not found in registry")
    return None


# Fixed version handling the generator issue

def _save_updated_registry_with_full_preservation(registry_sheet, df_existing: pd.DataFrame):
    """Save the updated registry to the Excel file while fully preserving styling, sheet order and tables.

    This approach is based on creating a copy of the workbook and then updating just the data,
    rather than recreating sheets from scratch.

    Args:
        registry_sheet: Excel worksheet object to update
        df_existing: DataFrame containing updated registry data
    """
    logger.debug("Saving updated registry to Excel file with full preservation")

    # Verify DataFrame isn't empty before proceeding
    if df_existing.empty:
        logger.error("Cannot save an empty DataFrame to Excel")
        raise ValueError("DataFrame is empty, no data to save")

    try:
        # Get the parent workbook and sheet info
        workbook = registry_sheet.parent
        sheet_name = registry_sheet.title

        # Get original sheet names and their order
        original_sheet_order = workbook.sheetnames
        logger.debug(f"Original sheet order: {original_sheet_order}")

        # Approach: Instead of creating a new sheet, we'll update the existing sheet's data
        # while preserving its formatting, tables, filters, etc.

        # Get the number of rows and columns in our data
        num_rows = len(df_existing) + 1  # +1 for header
        num_cols = len(df_existing.columns)

        # Clear the existing content in the relevant area
        logger.debug(f"Clearing content from sheet {sheet_name} for data size: {num_rows}x{num_cols}")

        # Instead of cell-by-cell clearing, which can be slow, we'll just overwrite with new data
        for row_idx in range(1, registry_sheet.max_row + 1):
            if row_idx <= num_rows:
                continue  # We'll overwrite these rows with new data

            # Delete rows beyond our data set
            for col_idx in range(1, registry_sheet.max_column + 1):
                cell = registry_sheet.cell(row=row_idx, column=col_idx)
                cell.value = None

        # Write the header row (row 1)
        header_row = list(df_existing.columns)
        logger.debug(f"Writing header row: {header_row}")
        for col_idx, header in enumerate(header_row, start=1):
            cell = registry_sheet.cell(row=1, column=col_idx)
            cell.value = header

        # Write data rows
        logger.debug(f"Writing {len(df_existing)} data rows")
        for row_idx, (_, row) in enumerate(df_existing.iterrows(), start=2):  # Start from row 2 (after header)
            for col_idx, col_name in enumerate(header_row, start=1):
                cell = registry_sheet.cell(row=row_idx, column=col_idx)
                value = row.get(col_name)

                # Handle datetime formatting consistently
                if isinstance(value, datetime):
                    cell.value = value
                    # Use Excel's built-in date format
                    cell.number_format = 'DD.MM.YYYY'
                else:
                    cell.value = value

        # Handle any cells beyond our data columns (clear them)
        for col_idx in range(num_cols + 1, registry_sheet.max_column + 1):
            for row_idx in range(1, num_rows + 1):
                cell = registry_sheet.cell(row=row_idx, column=col_idx)
                cell.value = None

        # Update any table references if needed
        if hasattr(registry_sheet, '_tables') and registry_sheet._tables:
            logger.debug("Updating table references")
            try:
                if isinstance(registry_sheet._tables, list):
                    for table in registry_sheet._tables:
                        if hasattr(table, 'ref'):
                            # Parse the current reference
                            ref_parts = table.ref.split(':')
                            if len(ref_parts) == 2:
                                start_ref, _ = ref_parts
                                # Extract column letters from end reference
                                col_letter = openpyxl.utils.get_column_letter(num_cols)
                                # Create new end reference with current row count
                                new_end_ref = f"{col_letter}{num_rows}"
                                new_ref = f"{start_ref}:{new_end_ref}"
                                logger.debug(f"Updating table reference from {table.ref} to {new_ref}")
                                table.ref = new_ref
                elif isinstance(registry_sheet._tables, dict):
                    for table_name, table in registry_sheet._tables.items():
                        if hasattr(table, 'ref'):
                            # Similar handling as above
                            ref_parts = table.ref.split(':')
                            if len(ref_parts) == 2:
                                start_ref, _ = ref_parts
                                col_letter = openpyxl.utils.get_column_letter(num_cols)
                                new_end_ref = f"{col_letter}{num_rows}"
                                new_ref = f"{start_ref}:{new_end_ref}"
                                logger.debug(f"Updating table reference from {table.ref} to {new_ref}")
                                table.ref = new_ref
            except Exception as e:
                logger.warning(f"Error updating table references: {e}")

        # Save the workbook
        logger.debug(f"Saving workbook to: {OUTPUT_FILE}")
        try:
            # Before saving, ensure sheets are in the original order
            # openpyxl manages sheets as OrderedDict, so order matters
            # To reorder, we need to compare current order with original and move sheets as needed
            current_order = workbook.sheetnames
            if current_order != original_sheet_order:
                logger.debug(f"Reordering sheets from {current_order} to {original_sheet_order}")

                # For each sheet in the original order, move it to the right position
                for i, sheet_name in enumerate(original_sheet_order):
                    if sheet_name in current_order:
                        current_index = current_order.index(sheet_name)
                        if current_index != i:
                            # Move sheet to the correct position
                            workbook.move_sheet(sheet_name, i)

                logger.debug(f"Sheet order after reordering: {workbook.sheetnames}")

            workbook.save(OUTPUT_FILE)
            logger.info(f"Workbook saved successfully to: {OUTPUT_FILE}")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}", exc_info=True)

            # If the full preservation save fails, try the simple version as fallback
            logger.warning("Falling back to simple save method")
            _save_updated_registry_simple(workbook, df_existing, sheet_name)

    except Exception as e:
        logger.error(f"Error in _save_updated_registry_with_full_preservation: {e}", exc_info=True)
        # Try the simple approach as a fallback
        try:
            logger.warning("Attempting fallback save method")
            _save_updated_registry_simple(registry_sheet.parent, df_existing, registry_sheet.title)
        except Exception as fallback_e:
            logger.error(f"Fallback save also failed: {fallback_e}", exc_info=True)
            raise


# Alternative implementation based on your previous project
def _save_updated_registry_pandas_approach(registry_sheet, df_existing: pd.DataFrame):
    """Save the updated registry using pandas ExcelWriter with style preservation.

    This approach is based on your previous project's method that worked well.

    Args:
        registry_sheet: Excel worksheet object to update
        df_existing: DataFrame containing updated registry data
    """
    logger.debug("Saving updated registry using pandas approach")

    # Verify DataFrame isn't empty before proceeding
    if df_existing.empty:
        logger.error("Cannot save an empty DataFrame to Excel")
        raise ValueError("DataFrame is empty, no data to save")

    try:
        # Get the parent workbook and sheet info
        workbook = registry_sheet.parent
        sheet_name = registry_sheet.title

        # Save a copy of the original file first
        original_file = os.path.join(os.path.dirname(OUTPUT_FILE), "original_copy.xlsx")
        logger.debug(f"Saving original file to: {original_file}")
        workbook.save(original_file)

        # Now, use pandas with the openpyxl engine to write the data
        # While preserving the existing workbook structure
        logger.debug(f"Writing {len(df_existing)} rows to {sheet_name}")

        # Create ExcelWriter with the mode='a' (append) to preserve existing sheets
        with pd.ExcelWriter(
                original_file,
                engine='openpyxl',
                mode='a',  # Append mode
                if_sheet_exists='replace',  # Replace specific sheet
                datetime_format='DD.MM.YYYY',
                date_format='DD.MM.YYYY'
        ) as writer:
            # Write the DataFrame to the specified sheet
            df_existing.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                float_format="%.2f"  # Format floats to 2 decimal places
            )

            # Get the workbook and sheet objects
            wb = writer.book
            ws = writer.sheets[sheet_name]

            # Apply auto-filter to the header row
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df_existing.columns))}{len(df_existing) + 1}"

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Now, copy the file back to the OUTPUT_FILE location
        import shutil
        logger.debug(f"Copying from {original_file} to {OUTPUT_FILE}")
        shutil.copy2(original_file, OUTPUT_FILE)

        # Clean up the temporary file
        try:
            os.remove(original_file)
        except:
            pass

        logger.info(f"Workbook saved successfully to: {OUTPUT_FILE}")

    except Exception as e:
        logger.error(f"Error in _save_updated_registry_pandas_approach: {e}", exc_info=True)
        # Try the simple approach as a fallback
        try:
            logger.warning("Attempting fallback save method")
            _save_updated_registry_simple(registry_sheet.parent, df_existing, registry_sheet.title)
        except Exception as fallback_e:
            logger.error(f"Fallback save also failed: {fallback_e}", exc_info=True)
            raise


# Third alternative using direct openpyxl copy and load
def _save_updated_registry_copy_approach(registry_sheet, df_existing: pd.DataFrame):
    """Save the updated registry by creating a full copy of the original workbook
    and then replacing just the data in the target sheet.

    This approach should preserve all styling, filters, tables, and sheet order.

    Args:
        registry_sheet: Excel worksheet object to update
        df_existing: DataFrame containing updated registry data
    """
    logger.debug("Saving updated registry using workbook copy approach")

    # Verify DataFrame isn't empty before proceeding
    if df_existing.empty:
        logger.error("Cannot save an empty DataFrame to Excel")
        raise ValueError("DataFrame is empty, no data to save")

    try:
        # Get the parent workbook and sheet info
        original_workbook = registry_sheet.parent
        original_file_path = original_workbook.path if hasattr(original_workbook, 'path') else None

        if not original_file_path:
            # If we can't get the original path, fall back to the simpler method
            logger.warning("Cannot determine original workbook path, falling back to simple save")
            _save_updated_registry_simple(original_workbook, df_existing, registry_sheet.title)
            return

        sheet_name = registry_sheet.title

        # Save the workbook to create a copy
        temp_file = os.path.join(os.path.dirname(OUTPUT_FILE), "temp_copy.xlsx")
        logger.debug(f"Saving copy to: {temp_file}")
        original_workbook.save(temp_file)

        # Close the original workbook to avoid file locks
        # This might be unnecessary but just to be safe
        try:
            original_workbook._archive.close()
        except:
            pass

        # Now load the copy and update just the data
        logger.debug(f"Loading copy and updating data in {sheet_name}")
        wb = openpyxl.load_workbook(temp_file)
        ws = wb[sheet_name]

        # Clear existing data (but keep formatting)
        for row in ws.iter_rows(min_row=2):  # Start from row 2 to preserve header
            for cell in row:
                cell.value = None

        # Write header row if needed (usually we'd keep the existing header)
        header_row = list(df_existing.columns)
        for col_idx, header in enumerate(header_row, start=1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value != header:  # Only update if different
                cell.value = header

        # Write data
        for row_idx, (_, row) in enumerate(df_existing.iterrows(), start=2):
            for col_idx, col_name in enumerate(header_row, start=1):
                value = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx)

                # Handle datetime formatting
                if isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'DD.MM.YYYY'
                else:
                    cell.value = value

        # Ensure auto-filter is applied
        max_row = len(df_existing) + 1  # +1 for header
        max_col = len(header_row)
        max_col_letter = openpyxl.utils.get_column_letter(max_col)

        ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

        # If there are tables, update their references
        if hasattr(ws, '_tables') and ws._tables:
            # Handle tables similar to previous implementations
            pass

        # Save the updated workbook
        logger.debug(f"Saving updated workbook to: {OUTPUT_FILE}")
        wb.save(OUTPUT_FILE)

        # Clean up
        try:
            os.remove(temp_file)
        except:
            pass

        logger.info(f"Workbook saved successfully to: {OUTPUT_FILE}")

    except Exception as e:
        logger.error(f"Error in _save_updated_registry_copy_approach: {e}", exc_info=True)
        # Try the simple approach as a fallback
        try:
            logger.warning("Attempting fallback save method")
            _save_updated_registry_simple(registry_sheet.parent, df_existing, registry_sheet.title)
        except Exception as fallback_e:
            logger.error(f"Fallback save also failed: {fallback_e}", exc_info=True)
            raise

# The simple save method remains the same
def _save_updated_registry_simple(workbook, df_existing: pd.DataFrame, sheet_name: str):
    """A simplified save method that doesn't try to preserve styling.
    Used as fallback if the styled version fails.

    Args:
        workbook: Excel workbook object
        df_existing: DataFrame containing updated registry data
        sheet_name: Name of the sheet to save to
    """
    logger.debug("Using simple save method")

    # Create a new workbook
    from openpyxl import Workbook
    new_workbook = Workbook()

    # Remove default sheet and create new one with same name
    if "Sheet" in new_workbook.sheetnames:
        new_workbook.remove(new_workbook["Sheet"])

    new_sheet = new_workbook.create_sheet(title=sheet_name)

    # Write header row
    header_row = list(df_existing.columns)
    new_sheet.append(header_row)

    # Apply bold to header
    from openpyxl.styles import Font
    for cell in new_sheet[1]:
        cell.font = Font(bold=True)

    # Write data rows
    for _, row in df_existing.iterrows():
        new_sheet.append([row.get(col) for col in header_row])

    # Auto-size columns
    for col in new_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        new_sheet.column_dimensions[column].width = adjusted_width

    # Save workbook
    new_workbook.save(OUTPUT_FILE)
    logger.info(f"Workbook saved with simple method to: {OUTPUT_FILE}")